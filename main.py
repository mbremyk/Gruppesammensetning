import openpyxl as opxl
from student import Student
from groupregister import GroupRegister
from app import App
from tkinter import StringVar

defaultmaxmembers = 5


def handlefilenamechange(filename: StringVar) -> None:
    """Method for handling changes in App.filename

    This method fetches data from en excel sheet with the filename specified in the StringVar filename,
    and creates a new GroupRegister with the data from the sheet
    If the StringVar is empty, nothing happens.
    The method is used as an event handler, but takes no event arguments

    :param filename: The path to the spreadsheet to read
    :type filename: StringVar

    :returns: None
    :rtype: None

    :raises IOException
    :raises InvalidFileException: if file is not supported by openpyxl
    """

    global groupregister
    if filename.get():
        groupregister = GroupRegister(app.spinmaxmembers.get())
        app.groupregister = groupregister
        wb = opxl.load_workbook(filename.get())
        ws = wb.active
        # Cuts off the first three cells of the first row, as they don't contain anything we want
        headers = [c.value[:100] for c in ws[1][3:]]
        # Swap the first and second header, as I want name first. Deal with it
        headers[0], headers[1] = headers[1], headers[0]
        app.headers = headers
        for row in ws.iter_rows(2):
            vals = []
            for cell in row:
                vals.append(cell.value)
            # vals[3:8] == [email, name, username, programming experience, preferred worktime, string of desired partners]
            # The first three cells don't contain valuable information
            student = Student(vals[3], vals[4], vals[5], vals[6], vals[7], vals[8])
            if not groupregister.getstudentbyemail(vals[3]):
                groupregister += student
            else:
                groupregister.updatestudent(student, email=vals[3], username=vals[5])


def createapp():
    """
    Creates an App object. This method did contain more, but it was moved inside the App class

    :return: App object
    :rtype: App
    """

    temp = App(title='Gruppesammensetning', geometry='1400x700')
    temp.filename.trace_add('write', lambda name, index, mode, filename=temp.filename: handlefilenamechange(filename))

    return temp


if __name__ == '__main__':
    app = createapp()
    app.mainloop()
